from openpyxl import Workbook, load_workbook


class Excel:
    def __init__(self, workbook_path):
        self.data = None
        self.colnum = None
        self.rownum = None
        self.sheet_name = None
        self.filename = None
        self.workbook_path = workbook_path
        self.wb = load_workbook(self.workbook_path)
        self.ws = self.wb.active

    def save(self):
        self.wb.save(f"{self.workbook_path}")

    def save_to(self, filename):
        self.filename = filename
        self.wb.save(f"{self.filename}")

    def getRowCount(self, sheet_name):
        self.sheet_name = sheet_name
        sheet = self.wb[self.sheet_name]
        return sheet.max_row

    def getColumnCount(self, sheet_name):
        self.sheet_name = sheet_name
        sheet = self.wb[sheet_name]
        return sheet.max_column

    def readData(self, sheet_name, rownum, colnum):
        self.rownum = rownum
        self.colnum = colnum
        self.sheet_name = sheet_name
        sheet = self.wb[self.sheet_name]
        return sheet.cell(row=self.rownum, column=self.colnum)

    def writeData(self, sheet_name, rownum, colnum, data):
        self.sheet_name = sheet_name
        self.rownum = rownum
        self.colnum = colnum
        self.data = data
        sheet = self.wb[self.sheet_name]
        sheet.cell(row=self.rownum, column=self.colnum).value = self.data
        # save file